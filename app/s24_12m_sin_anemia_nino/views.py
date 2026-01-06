# Standard library imports
import locale
import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List

# Third-party imports
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from django.views.generic import View
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import getpass


# Django imports
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import IntegerField,CharField
from django.db.models.functions import Cast, Substr
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.generic.base import TemplateView
import getpass

# Local imports
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo, Actualizacion
from .queries import obtener_velocimetro_s24_12m_sin_anemia_nino, obtener_grafico_mensual_s24_12m_sin_anemia_nino, obtener_variables_s24_12m_sin_anemia_nino, obtener_variables_detallado_s24_12m_sin_anemia_nino, obtener_grafico_por_redes_s24_12m_sin_anemia_nino
from .queries import obtener_grafico_por_microredes_s24_12m_sin_anemia_nino, obtener_grafico_por_establecimientos_s24_12m_sin_anemia_nino
from .queries import obtener_seguimiento_s24_12m_sin_anemia_nino

# Initialize logger and user model
logger = logging.getLogger(__name__)
User = get_user_model()

# Constants
VALID_YEARS = ['2024', '2025', '2026']
DEFAULT_YEAR = '2025'
GOBIERNO_REGIONAL = 'GOBIERNO REGIONAL'
DISA_JUNIN = 'JUNIN'

############################
## HELPER FUNCTIONS
############################

def _get_default_velocimetro_data() -> Dict[str, List]:
    """Retorna estructura por defecto para datos del velocímetro."""
    return {
        'numerador': [0],
        'denominador': [0],
        'avance': [0.0]
    }

def _extract_velocimetro_values(row: Dict[str, any]) -> tuple:
    """
    Extrae y valida valores del velocímetro desde una fila de datos.
    Args:
        row: Diccionario con datos de NUM, DEN, AVANCE
    Returns:
        Tupla (numerador, denominador, avance) con valores validados
    """
    numerador = row.get('NUM', 0)
    denominador = row.get('DEN', 0)
    avance = row.get('AVANCE', 0.0)
    
    # Asegurar que los valores no sean None
    numerador = numerador if numerador is not None else 0
    denominador = denominador if denominador is not None else 0
    avance = avance if avance is not None else 0.0
    
    return int(numerador), int(denominador), float(avance)

def _get_redes_queryset():
    """Obtiene queryset de redes de salud filtradas por región Junín."""
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(Descripcion_Sector=GOBIERNO_REGIONAL, Disa=DISA_JUNIN)
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )

def _get_provincias_queryset():
    """Obtiene queryset de provincias filtradas por sector gubernamental."""
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(Descripcion_Sector=GOBIERNO_REGIONAL)
        .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
        .values('Provincia', 'ubigueo_filtrado')
        .distinct()
        .order_by('Provincia')
    )


######################################
## PROCESOS DE COMPONENTES Y GRAFICOS 
######################################
## VELOCIMETRO
def process_velocimetro_s24_12m_sin_anemia_nino(resultados_velocimetro_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """
    Procesa los resultados del velocímetro para el formato del frontend.
    Args:
        resultados_velocimetro: Lista con un diccionario conteniendo NUM, DEN, AVANCE
    Returns:
        Diccionario con listas de numerador, denominador y avance
    """
    # Validar entrada
    if not resultados_velocimetro_s24_12m_sin_anemia_nino or len(resultados_velocimetro_s24_12m_sin_anemia_nino) == 0:
        logger.warning("Sin datos de velocímetro, usando valores por defecto")
        return _get_default_velocimetro_data()
    
    # Procesar el primer (y único) registro
    row = resultados_velocimetro_s24_12m_sin_anemia_nino[0]
    
    try:
        numerador, denominador, avance = _extract_velocimetro_values(row)
        
        logger.debug(f"Velocímetro procesado: Num={numerador}, Den={denominador}, Avance={avance}%")
        
        return {
            'numerador': [numerador],
            'denominador': [denominador],
            'avance': [avance]
        }
        
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Error procesando datos del velocímetro: {e}, Row: {row}")
        return _get_default_velocimetro_data()

## RESUMEN NUMERADOR Y DENOMINADOR 
def obtener_resumen_indicador_s24_12m_sin_anemia_nino(anio, mes_inicio, mes_fin, red, microred, establecimiento, provincia=None, distrito=None):
    """
    Obtiene un resumen detallado del indicador 
    """
    datos_base = obtener_velocimetro_s24_12m_sin_anemia_nino(anio, mes_inicio, mes_fin, red, microred, establecimiento, provincia, distrito)
    
    if not datos_base:
        return None
    
    resultado = datos_base[0]
    num = resultado.get('NUM', 0)
    den = resultado.get('DEN', 0)
    avance = resultado.get('AVANCE', 0.0)
    
    # Calcular métricas adicionales
    brecha = den - num
    porcentaje_brecha = (brecha / den * 100) if den > 0 else 0
    
    # Determinar clasificación
    if avance >= 82:
        clasificacion = "CUMPLE"
        color = "success"
        icono = "check-circle"
    elif avance >= 70:
        clasificacion = "EN PROCESO"
        color = "warning"
        icono = "clock"
    else:
        clasificacion = "EN RIESGO"
        color = "danger"
        icono = "exclamation-triangle"
    
    resumen = {
        'numerador': num,
        'denominador': den,
        'avance': round(avance, 2),
        'brecha': brecha,
        'porcentaje_brecha': round(porcentaje_brecha, 2),
        'clasificacion': clasificacion,
        'color': color,
        'icono': icono
    }
    
    return resumen

## GRAFICO MENSUALIZADO 
def process_avance_mensual_s24_12m_sin_anemia_nino(resultados_avance_mensual_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados del graficos"""
    data = {
        'num_1': [],
        'den_1': [],
        'cob_1': [],
        'num_2': [],
        'den_2': [],
        'cob_2': [],
        'num_3': [],
        'den_3': [],
        'cob_3': [],
        'num_4': [],
        'den_4': [],
        'cob_4': [],
        'num_5': [],
        'den_5': [],
        'cob_5': [],
        'num_6': [],
        'den_6': [],
        'cob_6': [],
        'num_7': [],
        'den_7': [],
        'cob_7': [],
        'num_8': [],
        'den_8': [],
        'cob_8': [],                
        'num_9': [],
        'den_9': [],
        'cob_9': [],
        'num_10': [],
        'den_10': [],
        'cob_10': [],
        'num_11': [],
        'den_11': [],
        'cob_11': [],
        'num_12': [],
        'den_12': [],
        'cob_12': [],
    }
    for index, row in enumerate(resultados_avance_mensual_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'num_1','den_1','cob_1','num_2','den_2','cob_2','num_3','den_3','cob_3','num_4','den_4','cob_4','num_5','den_5','cob_5','num_6','den_6','cob_6','num_7','den_7','cob_7','num_8','den_8','cob_8','num_9','den_9','cob_9','num_10','den_10','cob_10','num_11','den_11','cob_11','num_12','den_12','cob_12'}
            
            if not required_keys.issubset(row.keys()):
                raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")
            # Extraer cada valor, convirtiendo a float
            num_1_value = float(row.get('num_1', 0.0))
            den_1_value = float(row.get('den_1', 0.0))
            cob_1_value = float(row.get('cob_1', 0.0))
            num_2_value = float(row.get('num_2', 0.0))
            den_2_value = float(row.get('den_2', 0.0))
            cob_2_value = float(row.get('cob_2', 0.0))
            num_3_value = float(row.get('num_3', 0.0))
            den_3_value = float(row.get('den_3', 0.0))
            cob_3_value = float(row.get('cob_3', 0.0))
            num_4_value = float(row.get('num_4', 0.0))
            den_4_value = float(row.get('den_4', 0.0))
            cob_4_value = float(row.get('cob_4', 0.0))
            num_5_value = float(row.get('num_5', 0.0))
            den_5_value = float(row.get('den_5', 0.0))
            cob_5_value = float(row.get('cob_5', 0.0))
            num_6_value = float(row.get('num_6', 0.0))
            den_6_value = float(row.get('den_6', 0.0))
            cob_6_value = float(row.get('cob_6', 0.0))
            num_7_value = float(row.get('num_7', 0.0))
            den_7_value = float(row.get('den_7', 0.0))
            cob_7_value = float(row.get('cob_7', 0.0))
            num_8_value = float(row.get('num_8', 0.0))
            den_8_value = float(row.get('den_8', 0.0))
            cob_8_value = float(row.get('cob_8', 0.0))
            num_9_value = float(row.get('num_9', 0.0))
            den_9_value = float(row.get('den_9', 0.0))
            cob_9_value = float(row.get('cob_9', 0.0))
            num_10_value = float(row.get('num_10', 0.0))
            den_10_value = float(row.get('den_10', 0.0))
            cob_10_value = float(row.get('cob_10', 0.0))
            num_11_value = float(row.get('num_11', 0.0))
            den_11_value = float(row.get('den_11', 0.0))
            cob_11_value = float(row.get('cob_11', 0.0))
            num_12_value = float(row.get('num_12', 0.0))
            den_12_value = float(row.get('den_12', 0.0))
            cob_12_value = float(row.get('cob_12', 0.0))
            
            data['num_1'].append(num_1_value)
            data['den_1'].append(den_1_value)
            data['cob_1'].append(cob_1_value)
            data['num_2'].append(num_2_value)
            data['den_2'].append(den_2_value)
            data['cob_2'].append(cob_2_value)
            data['num_3'].append(num_3_value)
            data['den_3'].append(den_3_value)
            data['cob_3'].append(cob_3_value)
            data['num_4'].append(num_4_value)
            data['den_4'].append(den_4_value)
            data['cob_4'].append(cob_4_value)
            data['num_5'].append(num_5_value)
            data['den_5'].append(den_5_value)
            data['cob_5'].append(cob_5_value)
            data['num_6'].append(num_6_value)
            data['den_6'].append(den_6_value)
            data['cob_6'].append(cob_6_value)
            data['num_7'].append(num_7_value)
            data['den_7'].append(den_7_value)
            data['cob_7'].append(cob_7_value)
            data['num_8'].append(num_8_value)
            data['den_8'].append(den_8_value)
            data['cob_8'].append(cob_8_value)
            data['num_9'].append(num_9_value)
            data['den_9'].append(den_9_value)
            data['cob_9'].append(cob_9_value)
            data['num_10'].append(num_10_value)
            data['den_10'].append(den_10_value)
            data['cob_10'].append(cob_10_value)
            data['num_11'].append(num_11_value)
            data['den_11'].append(den_11_value)
            data['cob_11'].append(cob_11_value)
            data['num_12'].append(num_12_value)
            data['den_12'].append(den_12_value)
            data['cob_12'].append(cob_12_value)

        except Exception as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data

## GRAFICO VARIABLES 
def process_variables_s24_12m_sin_anemia_nino(resultados_variables_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados de las variables"""
    data = {
        'den_variable': [],
        'num_suple_inicio': [],
        'avance_suple_inicio': [],
        'num_suple_continuacion': [],
        'avance_suple_continuacion': [],
        'num_hb9m': [],
        'avance_hb9m': [],
        'num_hb12m': [],
        'avance_hb12m': []
    }
    for index, row in enumerate(resultados_variables_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'den_variable','num_suple_inicio','avance_suple_inicio','num_suple_continuacion','avance_suple_continuacion',
            'num_hb9m','avance_hb9m','num_hb12m','avance_hb12m'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores
            den_variable = row['den_variable']
            num_suple_inicio = row['num_suple_inicio']
            avance_suple_inicio = row['avance_suple_inicio']
            num_suple_continuacion = row['num_suple_continuacion']
            avance_suple_continuacion = row['avance_suple_continuacion']
            num_hb9m = row['num_hb9m']
            avance_hb9m = row['avance_hb9m']
            num_hb12m = row['num_hb12m']
            avance_hb12m = row['avance_hb12m']
            
            # Agrega los valores a la lista
            data['den_variable'].append(den_variable)
            data['num_suple_inicio'].append(num_suple_inicio)
            data['avance_suple_inicio'].append(avance_suple_inicio)
            data['num_suple_continuacion'].append(num_suple_continuacion)
            data['avance_suple_continuacion'].append(avance_suple_continuacion)
            data['num_hb9m'].append(num_hb9m)
            data['avance_hb9m'].append(avance_hb9m)
            data['num_hb12m'].append(num_hb12m)
            data['avance_hb12m'].append(avance_hb12m)
            
        except KeyError as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data

## TABLLA VARIABLES DETALLADOS
def process_variables_detallado_s24_12m_sin_anemia_nino(resultados_variables_detallado_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados de las variables detalladas
    NOTA: Usa prefijo 'detallado_' para las claves para NO sobrescribir los datos agregados"""
    data = {
        'd_anio': [],
        'd_mes': [],
        'd_codigo_red': [],
        'd_red': [],
        'd_codigo_microred': [],
        'd_microred': [],
        'd_codigo_unico': [],
        'd_id_establecimiento': [],
        'd_nombre_establecimiento': [],
        'd_ubigueo_establecimiento': [],
        'd_den_variable': [],
        'd_num_variable': [],
        'd_num_suple_inicio': [],
        'd_avance_suple_inicio': [],
        'd_num_suple_continuacion': [],
        'd_avance_suple_continuacion': [],
        'd_num_hb9m': [],
        'd_avance_hb9m': [],
        'd_num_hb12m': [],
        'd_avance_hb12m': []
    }
    for index, row in enumerate(resultados_variables_detallado_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'d_anio','d_mes','d_codigo_red','d_red','d_codigo_microred','d_microred','d_codigo_unico','d_id_establecimiento','d_nombre_establecimiento','d_ubigueo_establecimiento','d_den_variable','d_num_variable',
            'd_num_suple_inicio','d_avance_suple_inicio','d_num_suple_continuacion','d_avance_suple_continuacion','d_num_hb9m','d_avance_hb9m','d_num_hb12m','d_avance_hb12m'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores (las claves NO tienen el prefijo 'detallado_' en los datos)
            d_anio = row['d_anio']
            d_mes = row['d_mes']
            d_codigo_red = row['d_codigo_red']
            d_red = row['d_red']
            d_codigo_microred = row['d_codigo_microred']
            d_microred = row['d_microred']
            d_codigo_unico = row['d_codigo_unico']
            d_id_establecimiento = row['d_id_establecimiento']
            d_nombre_establecimiento = row['d_nombre_establecimiento']
            d_ubigueo_establecimiento = row['d_ubigueo_establecimiento']
            d_den_variable = row['d_den_variable']
            d_num_variable = row['d_num_variable']
            d_num_suple_inicio = row['d_num_suple_inicio']
            d_avance_suple_inicio = row['d_avance_suple_inicio']
            d_num_suple_continuacion = row['d_num_suple_continuacion']
            d_avance_suple_continuacion = row['d_avance_suple_continuacion']
            d_num_hb9m = row['d_num_hb9m']
            d_avance_hb9m = row['d_avance_hb9m']
            d_num_hb12m = row['d_num_hb12m']
            d_avance_hb12m = row['d_avance_hb12m']

            # Agrega los valores a la lista CON PREFIJO
            data['d_anio'].append(d_anio)
            data['d_mes'].append(d_mes)
            data['d_codigo_red'].append(d_codigo_red)
            data['d_red'].append(d_red)
            data['d_codigo_microred'].append(d_codigo_microred)
            data['d_microred'].append(d_microred)
            data['d_codigo_unico'].append(d_codigo_unico)
            data['d_id_establecimiento'].append(d_id_establecimiento)
            data['d_nombre_establecimiento'].append(d_nombre_establecimiento)
            data['d_ubigueo_establecimiento'].append(d_ubigueo_establecimiento)
            data['d_den_variable'].append(d_den_variable)
            data['d_num_variable'].append(d_num_variable)
            data['d_num_suple_inicio'].append(d_num_suple_inicio)
            data['d_avance_suple_inicio'].append(d_avance_suple_inicio)
            data['d_num_suple_continuacion'].append(d_num_suple_continuacion)
            data['d_avance_suple_continuacion'].append(d_avance_suple_continuacion)
            data['d_num_hb9m'].append(d_num_hb9m)
            data['d_avance_hb9m'].append(d_avance_hb9m)
            data['d_num_hb12m'].append(d_num_hb12m)
            data['d_avance_hb12m'].append(d_avance_hb12m)
            
        except KeyError as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data

## GRAFICO DE RANKING POR REDES
def process_grafico_por_redes_s24_12m_sin_anemia_nino(resultados_grafico_por_redes_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados del graficos por redes"""
    data = {
            'red_r': [],
            'den_r': [],
            'num_r': [],
            'avance_r': [],
            'brecha_r': [],
    }   
    for index, row in enumerate(resultados_grafico_por_redes_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'red_r','den_r','num_r','avance_r','brecha_r'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores
            red_r = row['red_r']
            den_r = row['den_r']
            num_r = row['num_r']
            avance_r = row['avance_r']
            brecha_r = row['brecha_r']
            
            # Agrega los valores a la lista
            data['red_r'].append(red_r)
            data['den_r'].append(den_r)
            data['num_r'].append(num_r)
            data['avance_r'].append(avance_r)
            data['brecha_r'].append(brecha_r)

        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")
    return data

## GRAFICO DE RANKING POR MICROREDES
def process_grafico_por_microredes_s24_12m_sin_anemia_nino(resultados_grafico_por_microredes_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados del graficos por microredes"""
    data = {
            'microred_mr': [],
            'den_mr': [],
            'num_mr': [],
            'avance_mr': [],
            'brecha_mr': [],
    }   
    for index, row in enumerate(resultados_grafico_por_microredes_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'microred_mr','den_mr','num_mr','avance_mr','brecha_mr'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores
            microred_mr = row['microred_mr']
            den_mr = row['den_mr']
            num_mr = row['num_mr']
            avance_mr = row['avance_mr']
            brecha_mr = row['brecha_mr']
            
            # Agrega los valores a la lista
            data['microred_mr'].append(microred_mr)
            data['den_mr'].append(den_mr)
            data['num_mr'].append(num_mr)
            data['avance_mr'].append(avance_mr)
            data['brecha_mr'].append(brecha_mr)

        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")
    return data

## GRAFICO DE RANKING POR ESTABLECIMIENTOS
def process_grafico_por_establecimientos_s24_12m_sin_anemia_nino(resultados_grafico_por_establecimientos_s24_12m_sin_anemia_nino: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados del graficos por establecimientos"""
    data = {
            'establecimiento_e': [],
            'den_e': [],
            'num_e': [],
            'avance_e': [],
            'brecha_e': [],
    }   
    for index, row in enumerate(resultados_grafico_por_establecimientos_s24_12m_sin_anemia_nino):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'establecimiento_e','den_e','num_e','avance_e','brecha_e'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores
            establecimiento_e = row['establecimiento_e']
            den_e = row['den_e']
            num_e = row['num_e']
            avance_e = row['avance_e']
            brecha_e = row['brecha_e']
            
            # Agrega los valores a la lista
            data['establecimiento_e'].append(establecimiento_e)
            data['den_e'].append(den_e)
            data['num_e'].append(num_e)
            data['avance_e'].append(avance_e)
            data['brecha_e'].append(brecha_e)

        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")
    return data

#######################
## PANTALLA PRINCIPAL
#######################

def index_s24_12m_sin_anemia_nino(request):
    """
    Vista principal para la pantalla de captación de gestantes.

    Maneja tanto la renderización inicial de la página como las peticiones AJAX
    para obtener datos del velocímetro según filtros aplicados.
    """
    # Obtener datos de actualización
    actualizacion = Actualizacion.objects.all()
    
    # Validar y obtener año
    anio = request.GET.get('anio', DEFAULT_YEAR)
    if anio not in VALID_YEARS:
        anio = DEFAULT_YEAR
    
    # Obtener parámetros de filtro
    mes_seleccionado_inicio = request.GET.get('mes_inicio')
    mes_seleccionado_fin = request.GET.get('mes_fin')
    provincia_seleccionada = request.GET.get('provincia_h')
    distrito_seleccionado = request.GET.get('distrito_h')
    red_seleccionada = request.GET.get('red_h')
    microred_seleccionada = request.GET.get('p_microredes_establec_h')
    establecimiento_seleccionado = request.GET.get('p_establecimiento_h')
    
    # Manejar peticiones AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Obtener datos del velocímetro con los filtros aplicados
            resultados_velocimetro_s24_12m_sin_anemia_nino = obtener_velocimetro_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,
                mes_fin=mes_seleccionado_fin,
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )

            # Obtener resumen del indicador
            resumen_s24_12m_sin_anemia_nino = obtener_resumen_indicador_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,
                mes_fin=mes_seleccionado_fin,
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )

            resultados_grafico_mensual_s24_12m_sin_anemia_nino = obtener_grafico_mensual_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,  # Siempre desde Enero,  # Usa el mes de inicio seleccionado
                mes_fin=mes_seleccionado_fin,      # Usa el mes de fin seleccionado
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )
            
            # Obtener datos del grafico mensualizado - SIEMPRE todos los meses del año
            # Los filtros de mes NO afectan el gráfico mensual, solo el velocímetro y resumen
            resultados_variables_s24_12m_sin_anemia_nino = obtener_variables_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,
                mes_fin=mes_seleccionado_fin, # Siempre hasta Diciembre
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )

            # Obtener datos de variables por trimestre - RESPETA los filtros de mes
            resultados_variables_detallado_s24_12m_sin_anemia_nino = obtener_variables_detallado_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,  # Usa el mes de inicio seleccionado
                mes_fin=mes_seleccionado_fin,        # Usa el mes de fin seleccionado
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )
            
            resultados_grafico_por_redes_s24_12m_sin_anemia_nino = obtener_grafico_por_redes_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,  # Usa el mes de inicio seleccionado
                mes_fin=mes_seleccionado_fin,        # Usa el mes de fin seleccionado
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )
            
            resultados_grafico_por_microredes_s24_12m_sin_anemia_nino = obtener_grafico_por_microredes_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,  # Usa el mes de inicio seleccionado
                mes_fin=mes_seleccionado_fin,        # Usa el mes de fin seleccionado
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )
            
            resultados_grafico_por_establecimientos_s24_12m_sin_anemia_nino = obtener_grafico_por_establecimientos_s24_12m_sin_anemia_nino(
                anio=anio,
                mes_inicio=mes_seleccionado_inicio,  # Usa el mes de inicio seleccionado
                mes_fin=mes_seleccionado_fin,        # Usa el mes de fin seleccionado
                red=red_seleccionada,
                microred=microred_seleccionada,
                establecimiento=establecimiento_seleccionado,
                provincia=provincia_seleccionada,
                distrito=distrito_seleccionado
            )

            # Procesar datos del velocímetro
            data = {
               **process_velocimetro_s24_12m_sin_anemia_nino(resultados_velocimetro_s24_12m_sin_anemia_nino),
               **process_avance_mensual_s24_12m_sin_anemia_nino(resultados_grafico_mensual_s24_12m_sin_anemia_nino),
               **process_variables_s24_12m_sin_anemia_nino(resultados_variables_s24_12m_sin_anemia_nino),
               **process_variables_detallado_s24_12m_sin_anemia_nino(resultados_variables_detallado_s24_12m_sin_anemia_nino),
               **process_grafico_por_redes_s24_12m_sin_anemia_nino(resultados_grafico_por_redes_s24_12m_sin_anemia_nino),
               **process_grafico_por_microredes_s24_12m_sin_anemia_nino(resultados_grafico_por_microredes_s24_12m_sin_anemia_nino),
               **process_grafico_por_establecimientos_s24_12m_sin_anemia_nino(resultados_grafico_por_establecimientos_s24_12m_sin_anemia_nino)
            }

            # Agregar datos del resumen a la respuesta
            if resumen_s24_12m_sin_anemia_nino:
                data['r_numerador_resumen'] = resumen_s24_12m_sin_anemia_nino['numerador']
                data['r_denominador_resumen'] = resumen_s24_12m_sin_anemia_nino['denominador']
                data['r_avance_resumen'] = resumen_s24_12m_sin_anemia_nino['avance']
                data['r_brecha'] = resumen_s24_12m_sin_anemia_nino['brecha']
                data['r_clasificacion'] = resumen_s24_12m_sin_anemia_nino['clasificacion']
                data['r_color'] = resumen_s24_12m_sin_anemia_nino['color']
                data['r_icono'] = resumen_s24_12m_sin_anemia_nino['icono']
            
            return JsonResponse(data)
            
        except Exception as e:
            logger.error(f"Error al obtener datos de captación de gestantes: {e}", exc_info=True)
            return JsonResponse(
                {'error': 'Error al obtener datos. Por favor, intente nuevamente.'},
                status=500
            )
    
    # Renderizado inicial de la página
    context = {
        'mes_seleccionado_inicio': mes_seleccionado_inicio,
        'mes_seleccionado_fin': mes_seleccionado_fin,
        'actualizacion': actualizacion,
        'provincia_seleccionada': provincia_seleccionada,
        'distrito_seleccionado': distrito_seleccionado,
        'provincias_h': _get_provincias_queryset(),
        'redes_h': _get_redes_queryset(),
    }
    
    return render(request, 's24_12m_sin_anemia_nino/index_s24_12m_sin_anemia_nino.html', context)


############################
## FILTROS HORIZONTAL
############################

def get_establecimientos_s24_12m_sin_anemia_nino_h(request, establecimiento_id):
    """
    Vista para renderizar la página de establecimientos con filtros horizontales.
    
    Args:
        request: HttpRequest
        establecimiento_id: ID del establecimiento (no usado actualmente, 
                           mantenido por compatibilidad con URL)
    
    Returns:
        Render del template con contexto de filtros
    """
    from .utils import build_filtro_context
    
    # Obtener el contexto completo de filtros usando la función reutilizable
    context = build_filtro_context(anio='2024')
    
    return render(request, 's24_12m_sin_anemia_nino/establecimientos_h.html', context)


def p_microredes_establec_s24_12m_sin_anemia_nino_h(request):
    """
    Vista parcial HTMX para cargar microredes según la red seleccionada.
    
    Args:
        request: HttpRequest con parámetro GET 'red_h'
    
    Returns:
        Render del partial con microredes filtradas
    """
    from .utils import get_microredes
    
    red_codigo = request.GET.get('red_h', '')
    
    # Usar la función reutilizable
    microredes = get_microredes(codigo_red=red_codigo) if red_codigo else []
    
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_microredes_establec_h.html', context)


def p_establecimientos_s24_12m_sin_anemia_nino_h(request):
    """
    Vista parcial HTMX para cargar establecimientos según microred o red seleccionada.
    Args:
        request: HttpRequest con parámetros GET:
                - 'p_microredes_establec_h': código de microred (opcional)
                - 'red_h': código de red (opcional)
    Returns:
        Render del partial con establecimientos filtrados
    """
    from .utils import get_establecimientos
    
    microred_codigo = request.GET.get('p_microredes_establec_h', '')
    red_codigo = request.GET.get('red_h', '')
    
    # Usar la función reutilizable con filtros dinámicos
    establecimientos = get_establecimientos(
        codigo_microred=microred_codigo if microred_codigo else None,
        codigo_red=red_codigo if red_codigo else None
    )
    
    # Convertir a lista para debug si es necesario
    establecimientos_list = list(establecimientos)
    
    context = {
        'establec': establecimientos_list
    }
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_establecimientos_h.html', context)


def p_distritos_s24_12m_sin_anemia_nino_h(request):
    """
    Vista parcial HTMX para cargar distritos según la provincia seleccionada.
    
    Args:
        request: HttpRequest con parámetro GET 'provincia'
    
    Returns:
        Render del partial con distritos filtrados
    """
    from .utils import get_distritos, get_provincias
    
    provincia_ubigueo = request.GET.get('provincia', '')
    
    # Obtener provincias para el contexto
    provincias = get_provincias()
    
    # Obtener distritos usando la función reutilizable
    distritos = get_distritos(ubigueo_provincia=provincia_ubigueo) if provincia_ubigueo else []
    
    context = {
        'distritos': distritos,
        'provincias_h': provincias
    }
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_distritos.html', context)


###########################################
## SEGUIMIENTO NOMINAL FILTROS
##########################################

###########################################
## FILTRO AMBITO DE SALUD
###########################################

## ============================================
# CONFIGURACIÓN BASE
# ============================================
FILTROS_BASE = {
    'Descripcion_Sector': 'GOBIERNO REGIONAL',
    'Disa': 'JUNIN'
}


FILTROS_BASE_ESTABLECIMIENTO = {
    'Descripcion_Sector': 'GOBIERNO REGIONAL',
    'Disa': 'JUNIN'
}

# ============================================
# HELPER FUNCTIONS - QUERIES REUTILIZABLES
# ============================================

def _get_redes_queryset():
    """
    Obtiene las redes de salud del gobierno regional de Junín.
    Returns: QuerySet con Codigo_Red, Red y codigo_red_filtrado
    """
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(**FILTROS_BASE_ESTABLECIMIENTO)
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Codigo_Red', 'Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )


def _get_meses_queryset(anio=None):
    """
    Obtiene los meses disponibles para los filtros.
    Args:
        anio: Año opcional para filtrar (None = todos los años)
    Returns: QuerySet con Mes y nro_mes
    """
    queryset = DimPeriodo.objects.all()
    
    if anio:
        queryset = queryset.filter(Anio=anio)
    
    return (
        queryset
        .annotate(nro_mes=Cast('NroMes', IntegerField()))
        .values('Mes', 'nro_mes')
        .order_by('nro_mes')
        .distinct()
    )


def _get_microredes_queryset(codigo_red):
    """
    Obtiene las microredes según el código de red.
    Args:
        codigo_red: Código de la red (puede ser parcial para usar startswith)
    Returns: QuerySet con Codigo_MicroRed y MicroRed
    """
    if not codigo_red:
        return []
    
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(
            Codigo_Red__startswith=codigo_red,
            **FILTROS_BASE_ESTABLECIMIENTO
        )
        .values('Codigo_MicroRed', 'MicroRed')
        .distinct()
        .order_by('MicroRed')
    )


def _get_establecimientos_queryset(codigo_microred, codigo_red=None):
    """
    Obtiene los establecimientos según la microred.
    Args:
        codigo_microred: Código de la microred
        codigo_red: Código de la red (opcional, para filtro adicional)
    Returns: QuerySet con Codigo_Unico y Nombre_Establecimiento
    """
    if not codigo_microred:
        return []
    
    filtros = {
        'Codigo_MicroRed__startswith': codigo_microred,
        **FILTROS_BASE_ESTABLECIMIENTO
    }
    
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(**filtros)
        .values('Codigo_Unico', 'Nombre_Establecimiento')
        .distinct()
        .order_by('Nombre_Establecimiento')
    )


def _get_context_base_con_filtros(include_meses=True, anio_meses=None):
    """
    Genera el contexto base común para los formularios.
    Args:
        include_meses: Si incluir los meses en el contexto
        anio_meses: Año para filtrar meses (None = todos)
    Returns: dict con redes y opcionalmente meses
    """
    context = {
        'redes': _get_redes_queryset(),
    }
    
    if include_meses:
        meses = _get_meses_queryset(anio_meses)
        context.update({
            'mes_inicio': meses,
            'mes_fin': meses,
        })
    
    return context


# ============================================
# VISTAS PRINCIPALES - FORMULARIOS
# ============================================

def get_redes_s24_12m_sin_anemia_nino(request, redes_id):
    """
    Renderiza el formulario de reportes por REDES.
    """
    context = _get_context_base_con_filtros(include_meses=True)
    return render(request, 's24_12m_sin_anemia_nino/components/salud/redes.html', context)


def get_microredes_s24_12m_sin_anemia_nino(request, microredes_id):
    """
    Renderiza el formulario de reportes por MICROREDES.
    """
    context = _get_context_base_con_filtros(include_meses=True, anio_meses='2024')
    return render(request, 's24_12m_sin_anemia_nino/components/salud/microredes.html', context)


def get_establecimientos_s24_12m_sin_anemia_nino(request, establecimiento_id):
    """Renderiza el formulario de reportes por ESTABLECIMIENTO."""
    context = {
        'redes': _get_redes_queryset(),
        'mes_inicio': _get_meses_queryset(),
        'mes_fin': _get_meses_queryset(),
    }
    return render(request, 's24_12m_sin_anemia_nino/components/salud/establecimientos.html', context)


# ============================================
# VISTAS PARTIALS - HTMX
# ============================================

def p_microredes_s24_12m_sin_anemia_nino(request):
    """
    Partial HTMX: Carga microredes según la red seleccionada.
    Usado en el formulario de MICROREDES (sin encadenamiento).
    
    GET params:
        - red: Código de la red seleccionada
    """
    red = request.GET.get('red', '').strip()
    microredes = list(_get_microredes_queryset(red))
    
    context = {
        'microredes': microredes,
        'red': red,
    }
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_microredes.html', context)


def p_microredes_establec_s24_12m_sin_anemia_nino(request):
    """
    HTMX Partial: Carga microredes según la red seleccionada.
    Se encadena con el select de establecimientos.
    """
    red = request.GET.get('red', '').strip()
    
    # Debug
    print(f"[p_microredes_establec] RED recibida: '{red}'")
    
    microredes = []
    if red:
        microredes = list(
            MAESTRO_HIS_ESTABLECIMIENTO.objects
            .filter(Codigo_Red__startswith=red, **FILTROS_BASE)
            .values('Codigo_MicroRed', 'MicroRed')
            .distinct()
            .order_by('MicroRed')
        )
        print(f"[p_microredes_establec] Microredes encontradas: {len(microredes)}")
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_microredes_establec.html', {
        'microredes': microredes,
        'red': red,
    })

# ============================================
# PARTIAL: ESTABLECIMIENTOS
# ============================================
def p_establecimientos_s24_12m_sin_anemia_nino(request):
    """
    HTMX Partial: Carga establecimientos según la microred seleccionada.
    """
    microred = request.GET.get('microred', '').strip()
    red = request.GET.get('red', '').strip()
    
    # Debug
    print(f"[p_establecimientos] MICRORED: '{microred}', RED: '{red}'")
    
    establecimientos = []
    if microred:
        filtros = {'Codigo_MicroRed': microred, **FILTROS_BASE}
        if red:
            filtros['Codigo_Red__startswith'] = red
        
        establecimientos = list(
            MAESTRO_HIS_ESTABLECIMIENTO.objects
            .filter(**filtros)
            .values('Codigo_Unico', 'Nombre_Establecimiento')
            .distinct()
            .order_by('Nombre_Establecimiento')
        )
        print(f"[p_establecimientos] Establecimientos encontrados: {len(establecimientos)}")
    
    return render(request, 's24_12m_sin_anemia_nino/partials/p_establecimientos.html', {
        'establecimientos': establecimientos,
    })
######################---------------------------
## FILTRO AMBITO DE MUNICIPIO
######################-------------------------------

## SEGUIMIENTO POR PROVINCIA
def get_provincias_s24_12m_sin_anemia_nino(request, provincia_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    )
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
            }
    
    return render(request, 's24_12m_sin_anemia_nino/components/municipio/provincias.html', context)

## SEGUIMIENTO POR DISTRITOS
def get_distritos_s24_12m_sin_anemia_nino(request, distrito_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 's24_12m_sin_anemia_nino/components/municipio/distritos.html', context)


def p_distrito_s24_12m_sin_anemia_nino(request):
    provincia_param = request.GET.get('provincia', '')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 's24_12m_sin_anemia_nino/partials/p_distritos.html', context)


########################################
## SEGUIMIENTO REPORTE EXCEL 
#######################################

# ============================================================================
# CONSTANTES Y CONFIGURACIONES
# ============================================================================

# Estilos de colores
COLORS = {
    'cyan': 'FF60D7E0',
    'orange': 'FFE0A960',
    'gray': 'FFD3D3D3',
    'green': 'FF60E0B3',
    'yellow': 'FFFFEB3B',
    'blue': 'FF60A2E0',
    'green_2': 'FF60E07E',
    'celeste': 'FF87CEEB',
    'morado_claro': 'FFE9D8FF',
    'plomo_claro': 'FFEDEDED',
    'azul_claro': 'FFD8EFFA',
    'naranja_claro': 'FFFFEBD8',
    'verde_claro': 'FFBDF7BD',
    'white': 'FFFFFF',
    'red': 'FF0000',
    'green_font': '00B050',
    'black': '000000',
    'dark_green': 'FF2E7D32',
}

# Anchos de columnas
COLUMN_WIDTHS = {
    'A': 1, 'B': 9, 'C': 9, 'D': 20, 'E': 9, 'F': 9, 'G': 7, 'H': 34,
    'I': 7, 'J': 9, 'K': 5, 'L': 9, 'M':5, 'N': 4, 'O': 9, 'P': 5,'Q': 9,
    'R': 5,'S': 10, 'T': 10, 'U':10, 'V':5, 'W':10, 'X':5, 'Y':10, 'Z':5,
    'AA':10,'AB':5,'AC':10,'AD':5,'AE':10, 'AF':5,'AG':10,'AH':5,'AI':10,'AJ':5,
    'AK':10,'AL':5,'AM':10,'AN':5,'AO':10, 'AP':10,'AQ':8,'AR':5,'AS':10,'AT':8,'AU':5,
    'AV':10,'AW':10,'AX':8,'AY':5,'AZ':10, 'BA':8,'BB':5,
    'BC':10,'BD':10,'BE':10,'BF':10,'BG':20,'BH':10,'BI':20,'BJ':10,'BK':25
}

# Alturas de filas
ROW_HEIGHTS = {1: 14, 2: 14, 3: 12, 4: 25, 5: 27, 6: 39, 7: 36, 8: 30}

# Configuración de cabeceras
HEADERS_CONFIG = [
    ('B9', 'TIPO DOC', 'cyan'),
    ('C9', 'NUM DOC', 'cyan'),
    ('D9', 'NOMBRE', 'cyan'),
    ('E9', 'DNI', 'cyan'),
    ('F9', 'CNV', 'cyan'),
    ('G9', 'COD UBI', 'cyan'),
    ('H9', 'CENTRO POBLADO','cyan'),
    ('I9', 'SEGURO','cyan'),
    ('J9', 'FEC NAC','cyan'),
    ('K9', 'SEXO','cyan'),
    ('L9', 'HB 6M','cyan'),        
    ('M9', 'VAL','cyan'),  
    ('N9', 'IND','cyan'),
    ('O9', 'DX A','cyan'),  
    ('P9', 'IND','cyan'),
    ('Q9', 'DX 12M','cyan'),
    ('R9', 'IND','cyan'),
    ('S9', 'DEN','cyan'),
    
    ('T9', 'INICIO SUP','yellow'),
    ('U9', '1° ENT','green_2'),
    ('V9', '1°','green_2'),
    ('W9', '2° ENT','green_2'),
    ('X9', '2°','green_2'),
    
    ('Y9', '1° MULT','green'),
    ('Z9', '1°','green'),
    ('AA9', '2° MULT','green'),
    ('AB9', '2°','green'),
    ('AC9', '3° MULT','green'),
    ('AD9', '3°','green'),
    ('AE9', '4° MULT','green'),
    ('AF9', '4°','green'),
    ('AG9', '5° MULT','green'),
    ('AH9', '5°','green'),
    ('AI9', '6° MULT','green'),
    ('AJ9', '6°','green'),

    ('AK9', '1°M + 2°FH','morado_claro'),
    ('AL9', 'MFH','morado_claro'),  
    ('AM9', '1°FH + 2°M','morado_claro'),  
    ('AN9', 'FHM','morado_claro'),  
    ('AO9', 'CONTINUO','yellow'),

    ('AP9', 'HB 9M FH','blue'),  
    ('AQ9', 'VAL','blue'),  
    ('AR9', 'FH','blue'),  
    ('AS9', 'HB 9M MMN','blue'),  
    ('AT9', 'VAL','blue'),  
    ('AU9', 'MMN','blue'),  
    ('AV9', 'HB 9M','yellow'),
    
    ('AW9', 'HB 12M FH','azul_claro'),  
    ('AX9', 'VAL','azul_claro'),  
    ('AY9', 'FH','azul_claro'),  
    ('AZ9', 'HB 12M MMN','azul_claro'),  
    ('BA9', 'VAL','azul_claro'),  
    ('BB9', 'MMN','azul_claro'),  
    ('BC9', 'HB 12M','yellow'),

    ('BD9', 'IND','blue'),
    ('BE9', 'MES','blue'),
    ('BF9', 'COD RED','orange'),
    ('BG9', 'RED','orange'),
    ('BH9', 'COD MICRO','orange'),
    ('BI9', 'MICRORED','orange'),
    ('BJ9', 'COD EESS','orange'),
    ('BK9', 'ESTABLECIMIENTO', 'orange')
]

# Celdas combinadas
MERGE_CELLS_CONFIG = [
    # Fila 5
    ('B5', 'S5'), ('T5', 'BC5'),
    # Fila 6
    ('B6', 'S6'), ('U6','X6'),('Y6','AJ6'),('AK6','AL6'), ('AM6','AN6'),('AO6','AO6'),('AP6','AV6'),('AW6','BC6'),
    # Fila 7
    ('B7','E7'),('F7','K7'),('L7','N7'),('O7','P7'),('Q7','R7'),('U7','X7'),('Y7','AJ7'),('AK7','AL7'),('AM7','AN7'),('AP7','AR7'),('AS7','AU7'),('AW7','AY7'),('AZ7','BB7'),
    # Fila 8
    ('B8','E8'),('F8','K8'),('L8','N8'),('O8','P8'),('Q8','R8'),('U8','V8'),('W8','X8'),('Y8','Z8'),('AA8','AB8'),('AC8','AD8'),('AE8','AF8'),('AG8','AH8'),('AI8','AJ8'),('AK8','AL8'),('AM8','AN8'),('AP8','AR8'),('AS8','AU8'),('AW8','AY8'),('AZ8','BB8'),
    # Columnas 
    ('S7','S8'),('AO7','AO8'),('AV7','AV8'),('BC7','BC8')
]

# ============================================================================
# CLASES DE UTILIDAD PARA ESTILOS
# ============================================================================

class ExcelStyleManager:
    """Gestor centralizado de estilos para Excel."""
    
    _fills_cache = {}
    _fonts_cache = {}
    _borders_cache = {}
    
    @classmethod
    def get_fill(cls, color_key):
        """Obtiene un PatternFill cacheado."""
        if color_key not in cls._fills_cache:
            color = COLORS.get(color_key, color_key)
            cls._fills_cache[color_key] = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )
        return cls._fills_cache[color_key]
    
    @classmethod
    def get_font(cls, name='Arial', size=8, bold=False, color='000000'):
        """Obtiene una Font cacheada."""
        key = (name, size, bold, color)
        if key not in cls._fonts_cache:
            cls._fonts_cache[key] = Font(name=name, size=size, bold=bold, color=color)
        return cls._fonts_cache[key]
    
    @classmethod
    def get_border(cls, color='A9A9A9', style='thin'):
        """Obtiene un Border cacheado."""
        key = (color, style)
        if key not in cls._borders_cache:
            side = Side(style=style, color=color)
            cls._borders_cache[key] = Border(
                left=side, right=side, top=side, bottom=side
            )
        return cls._borders_cache[key]
    
    @classmethod
    def get_alignment(cls, horizontal='center', vertical='center', wrap_text=False):
        """Obtiene un Alignment."""
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


# ============================================================================
# CLASE BASE PARA REPORTES
# ============================================================================

class BaseExcelReportView(LoginRequiredMixin, View):
    """Clase base para generar reportes Excel."""
    
    filename = "reporte.xlsx"
    sheet_name = "Datos"
    
    def get_query_params(self, request):
        """Extrae los parámetros de consulta. Debe ser implementado."""
        raise NotImplementedError("Subclases deben implementar get_query_params()")
    
    def get_data(self, params):
        """Obtiene los datos del reporte. Debe ser implementado."""
        raise NotImplementedError("Subclases deben implementar get_data()")
    
    def get_filename(self):
        """Retorna el nombre del archivo."""
        return self.filename
    
    def get(self, request, *args, **kwargs):
        """Maneja la petición GET y genera el Excel."""
        params = self.get_query_params(request)
        data = self.get_data(params)
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        
        fill_worksheet_optimized(ws, data, request.user)
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={self.get_filename()}"
        wb.save(response)
        
        return response


# ============================================================================
# VISTAS DE REPORTES
# ============================================================================

class Rpt12mSinAnemiaNino(BaseExcelReportView):
    """Reporte de captación de gestantes."""
    
    filename = "rpt_s24_12m_sin_anemia_nino.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
            'provincia': request.GET.get('provincia', ''),
            'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('p_microredes', ''),
            'establecimiento': request.GET.get('p_establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_s24_12m_sin_anemia_nino(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            params['provincia'], params['distrito'], params['red'],
            params['microredes'], params['establecimiento'], params['cumple']
        )


class Rpt12mSinAnemiaNinoMicroRed(BaseExcelReportView):
    """Reporte de población por microred."""
    
    filename = "rpt_s24_12m_sin_anemia_nino_microred.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
           # 'provincia': request.GET.get('provincia', ''),
           # 'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('p_microredes', ''),
           # 'establecimiento': request.GET.get('p_establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_s24_12m_sin_anemia_nino(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            '',  # provincia (vacío para microred)
            '',  # distrito (vacío para microred)
            params['red'],params['microredes'], 
              '',  # establecimiento (vacío para microred)
            params['cumple']
        )


class Rpt12mSinAnemiaNinoEstablec(BaseExcelReportView):
    """Reporte de población por establecimiento."""
    
    filename = "rpt_s24_12m_sin_anemia_nino_establecimiento.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
            'provincia': request.GET.get('provincia', ''),
            'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('microred', ''),
            'establecimiento': request.GET.get('establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_s24_12m_sin_anemia_nino(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            params['provincia'], params['distrito'], params['red'],
            params['microredes'], params['establecimiento'], params['cumple']
        )


# ============================================================================
# FUNCIONES DE FORMATEO
# ============================================================================

def fill_worksheet_optimized(ws, results, user=None):
    """Función optimizada para llenar la hoja de trabajo."""
    
    style_mgr = ExcelStyleManager
    
    # Configurar dimensiones
    _set_dimensions(ws)
    
    # Configurar agrupación de columnas
    _setup_column_grouping(ws)
    
    # Congelar paneles
    # ws.freeze_panes = 'S10'
    
    # Configurar celdas combinadas
    _setup_merged_cells(ws)
    
    # Aplicar estilos a secciones
    _style_header_sections(ws, style_mgr)
    
    # Configurar cabeceras de columnas
    _setup_column_headers(ws, style_mgr)
    
    # Agregar metadatos del reporte
    _add_report_metadata(ws, user, style_mgr)
    
    # Agregar títulos
    _add_titles(ws, style_mgr)
    
    # Escribir datos
    _write_data(ws, results, style_mgr)


def _set_dimensions(ws):
    """Configura las dimensiones de filas y columnas."""
    for row, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height
    
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _setup_column_grouping(ws):
    """Configura el agrupamiento de columnas."""
    #grouped_columns = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    #ws.column_dimensions.group('K', 'R', hidden=True)
    
    #for col in grouped_columns:
    #    ws.column_dimensions[col].width = COLUMN_WIDTHS.get(col, 10)
    #    ws.column_dimensions[col].outline_level = 1


def _setup_merged_cells(ws):
    """Configura las celdas combinadas."""
    for start, end in MERGE_CELLS_CONFIG:
        ws.merge_cells(f'{start}:{end}')


def _style_header_sections(ws, style_mgr):
    """Aplica estilos a las secciones de cabecera."""
    
    border_negro = style_mgr.get_border('000000')
    
    # Configuración de secciones con sus textos y estilos
    # NOTA: Solo incluir la celda INICIAL de cada rango combinado (MergedCell)
    # Las celdas dentro de un rango combinado que no son la inicial son read-only
    sections_config = {
        # Fila 5: B5:M5 y N5:AA5 son rangos combinados
        'B5': ('META (DENOMINADOR)', 'gray', 10, True),
        'T5': ('N° de niñas y niños del denominador, con al menos dos (02) entregas de suplementación preventiva con hierro y dos (02) dosajes de hemoglobina', 'naranja_claro', 10, True),
        # Fila 6: B6:M6, N6:S6, T6:Y6, Z6:AA6 son rangos combinados
        'B6': ('N° de niñas y niños de 12 meses de edad, registrados en el padrón nominal con DNI o CNV en línea, y sin anemia, en el mes de medición', 'gray', 10, True),
        'T6': ('1° Entrega de Hierro', 'gray', 8, True),
        'U6': ('Sulfato ferroso (SF) o Polimaltosado (HP)', 'gray', 10, True),
        'Y6': (' Multimicronutriente y continua sólo con Multimicronutriente (MMN)', 'gray', 10, True),
        'AK6': ('1° MMN + 2° SF/HP', 'gray', 10, True),
        'AM6': ('1° SF/HP + 2° MMN', 'gray', 10, True),
        'AP6': ('Entrega de dosaje de hemoglobina a los 3 meses de iniciada la suplementación con hierro', 'gray', 10, True),
        'AW6': ('Entrega de dosaje de hemoglobina a los 6 meses de iniciada la suplementación de hierro', 'gray', 10, True),


        # Fila 7: B7:D7, E7:H7, I7:J7, N7:P7, Q7:R7, T7:V7, W7:X7 son rangos combinados
        # NOTA: K7, L7 NO están combinadas (son celdas individuales)
        'B7': ('Padrón nominal de Niñas y niños de 12 meses 29 días de edad (394 días)', 'plomo_claro', 7, True),
        'F7': ('Se excluye las atenciones de telemedicina: 99499 y los registros que cuenten con códigos 99199.17 + LAB TA, en la primera entrega de hierro', 'plomo_claro', 7, False),
        'L7': ('Primer dosaje a los seis (06) meses de edad (se busca el dato entre 170-209 días)', 'plomo_claro', 7, False),
        'O7': ('Dx de anemia definitivo registrado a los seis (06) meses de edad', 'plomo_claro', 7, False),
        'Q7': ('Se excluye los Dx de Anemia (se busca el dato entre 210 - 364 días)', 'plomo_claro', 7, False),
        'T7': ('Cuentan con 1 frasco de Hierro el mismo dia del Dosaje', 'plomo_claro', 7, False),
        'U7': ('Sulfato ferroso o Polimaltosado la entrega tiene un intervalo de 25 a 70 días entre cada entrega', 'plomo_claro', 7, False),
        'Y7': ('Si la primera entrega es, las entregas deberán ser en un número de seis (06) con un intervalo de 25 a 35 entre cada entrega', 'plomo_claro', 7, False),
        'AK7': ('Intervalo entre cada entrega deberá ser de 25 a 70 días', 'plomo_claro', 7, False),
        'AM7': ('Cada entrega deberá ser de 25 a 35 días', 'plomo_claro', 7, False),
        'AP7': ('Se busca el dato entre 85 a 119 días SF/HP', 'plomo_claro', 7, False),
        'AS7': ('Se busca el dato entre 85 a 119 días MMN', 'plomo_claro', 7, False),
        'AW7': ('Se busca el dato entre 170 a 209 días MMN', 'plomo_claro', 7, False), 
        'AZ7': ('Se busca el dato entre 170 a 209 días SF/HP', 'plomo_claro', 7, False),
        # Fila 8: B8:D8, E8:H8, I8:J8, N8:P8, Q8:R8, T8:V8, W8:X8 son rangos combinados
        # NOTA: K8, L8, M8, S8, Y8, Z8, AA8 NO están combinadas
        'B8': ('CODIGO HIS MINSA', 'azul_claro', 7, True),
        'F8': ('', 'azul_claro', 7, True),
        'L8': ('DX = 85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),      
        'O8': ('DX = D509 ó D649 + TD=D', 'azul_claro', 7, False),
        'Q8': ('DX = D509 ó D649 + TD=D', 'azul_claro', 7, False),
        'T8': ('DX = 99199.17 ó 99199.19', 'azul_claro', 7, False),
        'U8': ('DX = 99199.17', 'azul_claro', 7, False),
        'W8': ('DX = 99199.17', 'azul_claro', 7, False),
        'Y8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AA8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AC8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AE8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AG8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AI8': ('DX = 99199.19', 'azul_claro', 7, False),
        'AK8': ('DX = 99199.17 + DX=99199.19', 'azul_claro', 7, False),
        'AM8': ('DX = 99199.19 + DX=99199.17', 'azul_claro', 7, False),
        'AP8': ('DX = 85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),
        'AS8': ('DX = 85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),
        'AW8': ('DX = 85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),
        'AZ8': ('DX = 85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),
    
    }
    
    for cell_ref, (text, fill_color, font_size, bold) in sections_config.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(wrap_text=True)
        cell.font = style_mgr.get_font(size=font_size, bold=bold)
        cell.fill = style_mgr.get_fill(fill_color)
        cell.border = border_negro
    
    # Aplicar bordes a las filas de cabecera
    _apply_row_borders(ws, [5, 6, 7, 8], 'B', 'BC', border_negro)


def _apply_row_borders(ws, rows, start_col, end_col, border):
    """Aplica bordes a rangos de celdas."""
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    
    for row in rows:
        for col in range(start_idx, end_idx + 1):
            ws.cell(row=row, column=col).border = border


def _setup_column_headers(ws, style_mgr):
    """Configura las cabeceras de columnas."""
    
    border = style_mgr.get_border('00B0F0')
    
    for cell_ref, text, fill_color in HEADERS_CONFIG:
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(wrap_text=True)
        cell.font = style_mgr.get_font(size=8, bold=True)
        cell.fill = style_mgr.get_fill(fill_color)
        cell.border = border


def _add_report_metadata(ws, user, style_mgr):
    """Agrega metadatos del reporte (fecha, hora, usuario)."""
    
    fecha_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = user.username if user else getpass.getuser()
    
    metadata = [
        ('Q1', 'Fecha y Hora:', 'R1', fecha_hora),
        ('Q2', 'Usuario:', 'R2', nombre_usuario),
    ]
    
    etiqueta_font = style_mgr.get_font(size=8)
    
    for label_ref, label, value_ref, value in metadata:
        ws[label_ref].value = label
        ws[label_ref].font = etiqueta_font
        ws[label_ref].alignment = style_mgr.get_alignment(horizontal='right')
        
        ws[value_ref].value = value
        ws[value_ref].font = etiqueta_font
        ws[value_ref].alignment = style_mgr.get_alignment(horizontal='left')


def _add_titles(ws, style_mgr):
    """Agrega los títulos del reporte."""
    
    titles = [
        ('B1', 'OFICINA DE TECNOLOGIAS DE LA INFORMACION', 7, True, '000000'),
        ('B2', 'DIRECCION REGIONAL DE SALUD JUNIN', 7, True, '000000'),
        ('B3', 'El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.', 7, True, '0000CC'),
        ('B4', 'SEGUIMIENTO NOMINAL:SI-02.04: PORCENTAJE DE NIÑAS Y NIÑOS DE DOCE (12) MESES DE EDAD DEL DEPARTAMENTO, SIN ANEMIA, QUE RECIBEN SUPLEMENTACIÓN PREVENTIVA CON HIERRO Y DOSAJES DE HEMOGLOBINA', 12, True, '000000'),
    ]
    
    for cell_ref, text, size, bold, color in titles:
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(horizontal='left')
        cell.font = style_mgr.get_font(size=size, bold=bold, color=color)


def _write_data(ws, results, style_mgr):
    """Escribe los datos en la hoja de trabajo."""
    
    border = style_mgr.get_border()
    check_mark = '✓'
    x_mark = '✗'
    
    # Columnas con alineación izquierda
    left_align_cols = {8, 61, 63}
    # Columnas con check/x marks
    check_cols = {14,16,18,22,24,26,28,30,32,34,36,38,40,44,47,51,54}
    # Columnas de sub-indicadores
    sub_indicator_cols = {19,20,41,48,55}
    
    for row_idx, record in enumerate(results, start=10):
        for col_idx, value in enumerate(record.values(), start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Determinar alineación
            if col_idx in left_align_cols:
                cell.alignment = style_mgr.get_alignment(horizontal='left')
            else:
                cell.alignment = style_mgr.get_alignment()
            
            # Aplicar formato según columna
            if col_idx == 56:  # Columna INDICADOR
                _format_indicator_cell(cell, value, style_mgr)
            elif col_idx in check_cols:
                _format_check_cell(cell, value, check_mark, x_mark, style_mgr)
            elif col_idx in sub_indicator_cols:
                _format_sub_indicator_cell(cell, value, style_mgr)
            else:
                cell.font = style_mgr.get_font(size=8)


def _format_indicator_cell(cell, value, style_mgr):
    """Formatea la celda de indicador."""
    if value == 0:
        cell.value = 'NO CUMPLE'
        cell.fill = style_mgr.get_fill('red')
        cell.font = style_mgr.get_font(size=8, bold=True, color='000000')
    elif value == 1:
        cell.value = 'CUMPLE'
        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')
        cell.font = style_mgr.get_font(size=8, bold=True, color='000000')
    else:
        cell.font = style_mgr.get_font(size=8, bold=True)


def _format_check_cell(cell, value, check_mark, x_mark, style_mgr):
    """Formatea celdas con check/x marks."""
    if value == 1:
        cell.value = check_mark
        cell.font = style_mgr.get_font(size=10, color='00B050')
    elif value == 0:
        cell.value = x_mark
        cell.font = style_mgr.get_font(size=10, color='FF0000')
    else:
        cell.font = style_mgr.get_font(size=8)


def _format_sub_indicator_cell(cell, value, style_mgr):
    """Formatea celdas de sub-indicadores."""
    if value == 0:
        cell.value = 'NO CUMPLE'
        cell.font = style_mgr.get_font(size=7, color='FF0000')
        cell.fill = style_mgr.get_fill('gray')
    elif value == 1:
        cell.value = 'CUMPLE'
        cell.font = style_mgr.get_font(size=7, color='00B050')
        cell.fill = style_mgr.get_fill('gray')  
    else:
        cell.font = style_mgr.get_font(size=7)