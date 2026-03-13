# Standard library imports
import locale
import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List

# Third-party imports
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from django.views.generic import View
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import getpass


# Django imports
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import IntegerField,CharField
from django.db.models.functions import Cast, Substr
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.generic.base import TemplateView
import getpass

# Local imports
import json
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo, Actualizacion
from .queries import obtener_velocimetro_dashboard, DEFAULT_VELOCIMETRO_DATA

# Initialize logger and user model
logger = logging.getLogger(__name__)
User = get_user_model()

# Constants
VALID_YEARS = ['2024', '2025', '2026']
DEFAULT_YEAR = '2026'
DEFAULT_MES = '1'  # Enero
GOBIERNO_REGIONAL = 'GOBIERNO REGIONAL'
DISA_JUNIN = 'JUNIN'

############################
## HELPER FUNCTIONS
############################

def _get_default_velocimetro_data() -> Dict[str, List]:
    """Retorna estructura por defecto para datos del velocímetro."""
    return {
        'numerador': [0],
        'denominador': [0],
        'avance': [0.0]
    }

def _extract_velocimetro_values(row: Dict[str, any]) -> tuple:
    """
    Extrae y valida valores del velocímetro desde una fila de datos.
    Args:
        row: Diccionario con datos de NUM, DEN, AVANCE
    Returns:
        Tupla (numerador, denominador, avance) con valores validados
    """
    numerador = row.get('NUM', 0)
    denominador = row.get('DEN', 0)
    avance = row.get('AVANCE', 0.0)
    
    # Asegurar que los valores no sean None
    numerador = numerador if numerador is not None else 0
    denominador = denominador if denominador is not None else 0
    avance = avance if avance is not None else 0.0
    
    return int(numerador), int(denominador), float(avance)

def _get_redes_queryset():
    """Obtiene queryset de redes de salud filtradas por región Junín."""
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(Descripcion_Sector=GOBIERNO_REGIONAL, Disa=DISA_JUNIN)
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )

def _get_provincias_queryset():
    """Obtiene queryset de provincias filtradas por sector gubernamental."""
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(Descripcion_Sector=GOBIERNO_REGIONAL)
        .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
        .values('Provincia', 'ubigueo_filtrado')
        .distinct()
        .order_by('Provincia')
    )


######################################
## PROCESOS DE COMPONENTES Y GRAFICOS 
######################################
## VELOCIMETRO

def process_velocimetro_dashboard(resultados_velocimetro_dashboard: List[Dict]) -> Dict[str, List]:
    """Procesa los resultados del velocímetro para el formato del frontend."""
    data = {
        'orden': [],
        'codigo': [],
        'codigo_red': [],
        'codigo_microred': [],
        'id_establecimiento': [],
        'red': [],
        'microred': [],
        'nombre_establecimiento': [],
        'den': [],
        'num': [],
        'avance': []
    }

    for index, row in enumerate(resultados_velocimetro_dashboard):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'orden','codigo','codigo_red','codigo_microred', 'id_establecimiento', 'red', 'microred', 'nombre_establecimiento', 'den', 'num', 'avance'}
            
            if not required_keys.issubset(row.keys()):
                raise KeyError(f"Falta una o más claves en la fila {index}: {required_keys - row.keys()}")
            
            # Extrae los valores (las claves NO tienen el prefijo 'detallado_' en los datos)
            orden = row['orden']
            codigo = row['codigo']
            codigo_red = row['codigo_red']
            codigo_microred = row['codigo_microred']
            id_establecimiento = row['id_establecimiento']
            red = row['red']
            microred = row['microred']
            nombre_establecimiento = row['nombre_establecimiento']
            den = row['den']
            num = row['num']
            avance = row['avance']
            
            # Agrega los valores a la lista CON PREFIJO
            data['orden'].append(orden)
            data['codigo'].append(codigo)
            data['codigo_red'].append(codigo_red)
            data['codigo_microred'].append(codigo_microred)
            data['id_establecimiento'].append(id_establecimiento)
            data['red'].append(red)
            data['microred'].append(microred)
            data['nombre_establecimiento'].append(nombre_establecimiento)
            data['den'].append(den)
            data['num'].append(num)
            data['avance'].append(avance)
            
        except KeyError as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data


def _obtener_datos_dashboard(anio, mes_inicio, mes_fin, red, microred, establecimiento):
    """
    Función auxiliar que obtiene y procesa los datos del dashboard.
    Retorna el diccionario procesado con las listas de datos.
    """
    try:
        resultados = obtener_velocimetro_dashboard(
            anio=anio,
            mes_inicio=mes_inicio,
            mes_fin=mes_fin,
            red=red,
            microred=microred,
            establecimiento=establecimiento
        )
        return process_velocimetro_dashboard(resultados)
    except Exception as e:
        logger.error(f"Error al obtener datos del dashboard: {e}", exc_info=True)
        return process_velocimetro_dashboard([DEFAULT_VELOCIMETRO_DATA])


#######################
## PANTALLA PRINCIPAL
#######################

@login_required(login_url='/signin/')
def index_dashboard(request):
    """
    Vista principal para la pantalla de captación de gestantes.

    Maneja tanto la renderización inicial de la página como las peticiones AJAX
    para obtener datos del velocímetro según filtros aplicados.
    """
    # Obtener datos de actualización
    actualizacion = Actualizacion.objects.all()
    
    # Validar y obtener año
    anio = request.GET.get('anio', DEFAULT_YEAR)
    if anio not in VALID_YEARS:
        anio = DEFAULT_YEAR
    
    # Obtener parámetros de filtro (con defaults para carga inicial)
    mes_seleccionado_inicio = request.GET.get('mes_inicio', DEFAULT_MES)
    mes_seleccionado_fin = request.GET.get('mes_fin', DEFAULT_MES)
    provincia_seleccionada = request.GET.get('provincia_h')
    distrito_seleccionado = request.GET.get('distrito_h')
    red_seleccionada = request.GET.get('red_h', '')
    microred_seleccionada = request.GET.get('p_microredes_establec_h', '')
    establecimiento_seleccionado = request.GET.get('p_establecimiento_h', '')
    
    # Obtener datos del dashboard (usado tanto para AJAX como para carga inicial)
    data = _obtener_datos_dashboard(
        anio=anio,
        mes_inicio=mes_seleccionado_inicio,
        mes_fin=mes_seleccionado_fin,
        red=red_seleccionada,
        microred=microred_seleccionada,
        establecimiento=establecimiento_seleccionado
    )
    
    # Manejar peticiones AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse(data)
    
    # Renderizado inicial de la página con datos precargados
    context = {
        'mes_seleccionado_inicio': mes_seleccionado_inicio,
        'mes_seleccionado_fin': mes_seleccionado_fin,
        'actualizacion': actualizacion,
        'provincia_seleccionada': provincia_seleccionada,
        'distrito_seleccionado': distrito_seleccionado,
        'provincias_h': _get_provincias_queryset(),
        'redes_h': _get_redes_queryset(),
        'anio': anio,
        # Datos iniciales del dashboard para renderizar en el template
        'data_inicial': data,
        'data_inicial_json': json.dumps(data),
    }
    
    return render(request, 'dashboard/index_dashboard.html', context)


############################
## FILTROS HORIZONTAL
############################

def get_establecimientos_dashboard_h(request, establecimiento_id):
    """
    Vista para renderizar la página de establecimientos con filtros horizontales.
    
    Args:
        request: HttpRequest
        establecimiento_id: ID del establecimiento (no usado actualmente, 
                           mantenido por compatibilidad con URL)
    
    Returns:
        Render del template con contexto de filtros
    """
    from .utils import build_filtro_context
    
    # Obtener el contexto completo de filtros usando la función reutilizable
    context = build_filtro_context(anio='2024')
    
    return render(request, 'dashboard/establecimientos_h.html', context)


def p_microredes_establec_dashboard_h(request):
    """
    Vista parcial HTMX para cargar microredes según la red seleccionada.
    
    Args:
        request: HttpRequest con parámetro GET 'red_h'
    
    Returns:
        Render del partial con microredes filtradas
    """
    from .utils import get_microredes
    
    red_codigo = request.GET.get('red_h', '')
    
    # Usar la función reutilizable
    microredes = get_microredes(codigo_red=red_codigo) if red_codigo else []
    
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    
    return render(request, 'dashboard/partials/p_microredes_establec_h.html', context)


def p_establecimientos_dashboard_h(request):
    """
    Vista parcial HTMX para cargar establecimientos según microred o red seleccionada.
    Args:
        request: HttpRequest con parámetros GET:
                - 'p_microredes_establec_h': código de microred (opcional)
                - 'red_h': código de red (opcional)
    Returns:
        Render del partial con establecimientos filtrados
    """
    from .utils import get_establecimientos
    
    microred_codigo = request.GET.get('p_microredes_establec_h', '')
    red_codigo = request.GET.get('red_h', '')
    
    # Usar la función reutilizable con filtros dinámicos
    establecimientos = get_establecimientos(
        codigo_microred=microred_codigo if microred_codigo else None,
        codigo_red=red_codigo if red_codigo else None
    )
    
    # Convertir a lista para debug si es necesario
    establecimientos_list = list(establecimientos)
    
    context = {
        'establec': establecimientos_list
    }
    
    return render(request, 'dashboard/partials/p_establecimientos_h.html', context)


def p_distritos_dashboard_h(request):
    """
    Vista parcial HTMX para cargar distritos según la provincia seleccionada.
    
    Args:
        request: HttpRequest con parámetro GET 'provincia'
    
    Returns:
        Render del partial con distritos filtrados
    """
    from .utils import get_distritos, get_provincias
    
    provincia_ubigueo = request.GET.get('provincia', '')
    
    # Obtener provincias para el contexto
    provincias = get_provincias()
    
    # Obtener distritos usando la función reutilizable
    distritos = get_distritos(ubigueo_provincia=provincia_ubigueo) if provincia_ubigueo else []
    
    context = {
        'distritos': distritos,
        'provincias_h': provincias
    }
    
    return render(request, 'dashboard/partials/p_distritos.html', context)


###########################################
## SEGUIMIENTO NOMINAL FILTROS
##########################################

###########################################
## FILTRO AMBITO DE SALUD
###########################################

## ============================================
# CONFIGURACIÓN BASE
# ============================================
FILTROS_BASE = {
    'Descripcion_Sector': 'GOBIERNO REGIONAL',
    'Disa': 'JUNIN'
}


FILTROS_BASE_ESTABLECIMIENTO = {
    'Descripcion_Sector': 'GOBIERNO REGIONAL',
    'Disa': 'JUNIN'
}

# ============================================
# HELPER FUNCTIONS - QUERIES REUTILIZABLES
# ============================================

def _get_redes_queryset():
    """
    Obtiene las redes de salud del gobierno regional de Junín.
    Returns: QuerySet con Codigo_Red, Red y codigo_red_filtrado
    """
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(**FILTROS_BASE_ESTABLECIMIENTO)
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Codigo_Red', 'Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )


def _get_meses_queryset(anio=None):
    """
    Obtiene los meses disponibles para los filtros.
    Args:
        anio: Año opcional para filtrar (None = todos los años)
    Returns: QuerySet con Mes y nro_mes
    """
    queryset = DimPeriodo.objects.all()
    
    if anio:
        queryset = queryset.filter(Anio=anio)
    
    return (
        queryset
        .annotate(nro_mes=Cast('NroMes', IntegerField()))
        .values('Mes', 'nro_mes')
        .order_by('nro_mes')
        .distinct()
    )


def _get_microredes_queryset(codigo_red):
    """
    Obtiene las microredes según el código de red.
    Args:
        codigo_red: Código de la red (puede ser parcial para usar startswith)
    Returns: QuerySet con Codigo_MicroRed y MicroRed
    """
    if not codigo_red:
        return []
    
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(
            Codigo_Red__startswith=codigo_red,
            **FILTROS_BASE_ESTABLECIMIENTO
        )
        .values('Codigo_MicroRed', 'MicroRed')
        .distinct()
        .order_by('MicroRed')
    )


def _get_establecimientos_queryset(codigo_microred, codigo_red=None):
    """
    Obtiene los establecimientos según la microred.
    Args:
        codigo_microred: Código de la microred
        codigo_red: Código de la red (opcional, para filtro adicional)
    Returns: QuerySet con Codigo_Unico y Nombre_Establecimiento
    """
    if not codigo_microred:
        return []
    
    filtros = {
        'Codigo_MicroRed__startswith': codigo_microred,
        **FILTROS_BASE_ESTABLECIMIENTO
    }
    
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    return (
        MAESTRO_HIS_ESTABLECIMIENTO.objects
        .filter(**filtros)
        .values('Codigo_Unico', 'Nombre_Establecimiento')
        .distinct()
        .order_by('Nombre_Establecimiento')
    )


def _get_context_base_con_filtros(include_meses=True, anio_meses=None):
    """
    Genera el contexto base común para los formularios.
    Args:
        include_meses: Si incluir los meses en el contexto
        anio_meses: Año para filtrar meses (None = todos)
    Returns: dict con redes y opcionalmente meses
    """
    context = {
        'redes': _get_redes_queryset(),
    }
    
    if include_meses:
        meses = _get_meses_queryset(anio_meses)
        context.update({
            'mes_inicio': meses,
            'mes_fin': meses,
        })
    
    return context


# ============================================
# VISTAS PRINCIPALES - FORMULARIOS
# ============================================

def get_redes_dashboard(request, redes_id):
    """
    Renderiza el formulario de reportes por REDES.
    """
    context = _get_context_base_con_filtros(include_meses=True)
    return render(request, 'dashboard/components/salud/redes.html', context)


def get_microredes_dashboard(request, microredes_id):
    """
    Renderiza el formulario de reportes por MICROREDES.
    """
    context = _get_context_base_con_filtros(include_meses=True, anio_meses='2024')
    return render(request, 'dashboard/components/salud/microredes.html', context)


def get_establecimientos_dashboard(request, establecimiento_id):
    """Renderiza el formulario de reportes por ESTABLECIMIENTO."""
    context = {
        'redes': _get_redes_queryset(),
        'mes_inicio': _get_meses_queryset(),
        'mes_fin': _get_meses_queryset(),
    }
    return render(request, 'dashboard/components/salud/establecimientos.html', context)


# ============================================
# VISTAS PARTIALS - HTMX
# ============================================

def p_microredes_dashboard(request):
    """
    Partial HTMX: Carga microredes según la red seleccionada.
    Usado en el formulario de MICROREDES (sin encadenamiento).
    
    GET params:
        - red: Código de la red seleccionada
    """
    red = request.GET.get('red', '').strip()
    microredes = list(_get_microredes_queryset(red))
    
    context = {
        'microredes': microredes,
        'red': red,
    }
    
    return render(request, 'dashboard/partials/p_microredes.html', context)


def p_microredes_establec_dashboard(request):
    """
    HTMX Partial: Carga microredes según la red seleccionada.
    Se encadena con el select de establecimientos.
    """
    red = request.GET.get('red', '').strip()
    
    # Debug
    print(f"[p_microredes_establec] RED recibida: '{red}'")
    
    microredes = []
    if red:
        microredes = list(
            MAESTRO_HIS_ESTABLECIMIENTO.objects
            .filter(Codigo_Red__startswith=red, **FILTROS_BASE)
            .values('Codigo_MicroRed', 'MicroRed')
            .distinct()
            .order_by('MicroRed')
        )
        print(f"[p_microredes_establec] Microredes encontradas: {len(microredes)}")
    
    return render(request, 'dashboard/partials/p_microredes_establec.html', {
        'microredes': microredes,
        'red': red,
    })

# ============================================
# PARTIAL: ESTABLECIMIENTOS
# ============================================
def p_establecimientos_dashboard(request):
    """
    HTMX Partial: Carga establecimientos según la microred seleccionada.
    """
    microred = request.GET.get('microred', '').strip()
    red = request.GET.get('red', '').strip()
    
    # Debug
    print(f"[p_establecimientos] MICRORED: '{microred}', RED: '{red}'")
    
    establecimientos = []
    if microred:
        filtros = {'Codigo_MicroRed': microred, **FILTROS_BASE}
        if red:
            filtros['Codigo_Red__startswith'] = red
        
        establecimientos = list(
            MAESTRO_HIS_ESTABLECIMIENTO.objects
            .filter(**filtros)
            .values('Codigo_Unico', 'Nombre_Establecimiento')
            .distinct()
            .order_by('Nombre_Establecimiento')
        )
        print(f"[p_establecimientos] Establecimientos encontrados: {len(establecimientos)}")
    
    return render(request, 'dashboard/partials/p_establecimientos.html', {
        'establecimientos': establecimientos,
    })
######################---------------------------
## FILTRO AMBITO DE MUNICIPIO
######################-------------------------------

## SEGUIMIENTO POR PROVINCIA
def get_provincias_dashboard(request, provincia_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    )
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
            }
    
    return render(request, 'dashboard/components/municipio/provincias.html', context)

## SEGUIMIENTO POR DISTRITOS
def get_distritos_dashboard(request, distrito_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'dashboard/components/municipio/distritos.html', context)


def p_distrito_dashboard(request):
    provincia_param = request.GET.get('provincia', '')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'dashboard/partials/p_distritos.html', context)


########################################
## SEGUIMIENTO REPORTE EXCEL 
#######################################

# ============================================================================
# CONSTANTES Y CONFIGURACIONES
# ============================================================================

# Estilos de colores
COLORS = {
    'cyan': 'FF60D7E0',
    'orange': 'FFE0A960',
    'gray': 'FFD3D3D3',
    'green': 'FF60E0B3',
    'yellow': 'FFFFEB3B',
    'blue': 'FF60A2E0',
    'green_2': 'FF60E07E',
    'celeste': 'FF87CEEB',
    'morado_claro': 'FFE9D8FF',
    'plomo_claro': 'FFEDEDED',
    'azul_claro': 'FFD8EFFA',
    'naranja_claro': 'FFFFEBD8',
    'verde_claro': 'FFBDF7BD',
    'white': 'FFFFFF',
    'red': 'FF0000',
    'green_font': '00B050',
    'black': '000000',
    'dark_green': 'FF2E7D32',
}

# Anchos de columnas
COLUMN_WIDTHS = {
    'A': 1, 'B': 9, 'C': 9, 'D': 9, 'E': 8, 'F': 9, 'G': 9, 'H': 9,'I':9, 'J':9, 'K':9,
    'L':9, 'M':9, 'N':5, 'O':9, 'P':5,'Q':9, 'R':5, 'S':9, 'T':5, 'U':9, 'V':5, 
    'W':9, 'X':9, 'Y':5, 'Z':9,'AA':5,'AB':9,'AC':5,'AD':9,'AE':5,'AF':9,'AG':5,'AH':5,'AI':5,
    'AJ':9,'AK':9,'AL':5,'AM':9,'AN':5,'AO':9,'AP':5,'AQ':9,'AR':5,'AS':9,'AT':5,
    'AU':12,'AV':5,'AW':9,'AX':20,'AY':9,'AZ':25,'BA':9,'BB':25
}

# Alturas de filas
ROW_HEIGHTS = {1: 14, 2: 14, 3: 12, 4: 25, 5: 27, 6: 39, 7: 36, 8: 30}

# Configuración de cabeceras
HEADERS_CONFIG = [
    ('B9', 'UBIGUEO', 'cyan'),
    ('C9', 'DNI', 'cyan'),
    ('D9', 'SEM GEST', 'cyan'),
    ('E9', 'INICIO GEST', 'cyan'),
    ('F9', 'SEM 14', 'cyan'),
    ('G9', 'SEM 28', 'cyan'),
    ('H9', 'SEM 33', 'cyan'),
    ('I9', 'SEM 37','cyan'),
    ('J9', 'PARTO','cyan'),
    ('K9', 'DEN','cyan'),
    ('L9', 'NUM EXAM','yellow'),
    ('M9', 'DOSAJE HB','yellow'),
    ('N9', 'HB','yellow'),
    ('O9', 'SIFILIS','yellow'),
    ('P9', 'S','yellow'),
    ('Q9', 'VIH','yellow'),
    ('R9', 'V','yellow'),
    ('S9', 'BACT','yellow'),
    ('T9', 'B','yellow'),
    ('U9', 'PERFIL OBS','yellow'),
    ('V9', 'PO','yellow'),
    ('W9', 'APN','green'),
    ('X9', '1° APN','green'),
    ('Y9', '1°','green'),
    ('Z9', '1° APN','green_2'),
    ('AA9', '1°','green_2'),
    ('AB9', '2° APN','green_2'),
    ('AC9', '2°','green_2'),
    ('AD9', '1° APN','green'),
    ('AE9', '1°','green'),
    ('AF9', '2° APN','green'),   
    ('AG9', '2°','green'),
    ('AH9', '3° APN','green'),
    ('AI9', '3°','green'),
    ('AJ9', 'ENTREGA','morado_claro'),
    ('AK9', '1° ENT','morado_claro'),
    ('AL9','1°','morado_claro'),
    ('AM9','2° ENT','morado_claro'),
    ('AN9','2°','morado_claro'),
    ('AO9','3° ENT','morado_claro'),
    ('AP9','3°','morado_claro'),
    ('AQ9','4° ENT','morado_claro'),
    ('AR9','4°','morado_claro'),
    ('AS9','5° ENT','morado_claro'),
    ('AT9','5°','morado_claro'), 
    ('AU9', 'IND','blue'),
    ('AV9', 'MES','blue'),
    ('AW9', 'COD RED','orange'),
    ('AX9', 'RED','orange'),
    ('AY9', 'COD MICRO','orange'),
    ('AZ9', 'MICRORED','orange'),
    ('BA9', 'COD EESS','orange'),
    ('BB9', 'ESTABLECIMIENTO', 'orange')
]

# Celdas combinadas
# NOTA: Solo se puede asignar valor a la celda top-left de cada rango combinado
# Las demás celdas del rango son read-only en openpyxl
MERGE_CELLS_CONFIG = [
    # Fila 5: B5-G5 para META, H5-K5 para título numerador (L5 queda libre para otro contenido)
    ('B5', 'K5'),('L5','AT5'),
    # Fila 6: B6-G6 para descripción denominador, H6-K6 para descripción numerador
    ('B6', 'K6'),('L6','V6'),('X6','Y6'),('Z6','AC6'),('AD6','AI6'),('AJ6','AT6'),
    # Fila 7: B7-C7 para descripción adolescentes, E7-F7 para descripción anemia
    ('B7','C7'),('M7','N7'),('O7','P7'),('Q7','R7'),('S7','T7'),('U7','V7'),
    ('X7','Y7'),('Z7','AA7'),('AB7','AC7'),('AD7','AE7'),('AF7','AG7'),('AH7','AI7'),
    ('AK7','AL7'),('AM7','AN7'),('AO7','AP7'),('AQ7','AR7'),('AS7','AT7'),
    # Fila 8: B8-C8 para descripción adolescentes, E8-F8 para descripción anemia
    ('B8','C8'),('M8','N8'),('O8','P8'),('Q8','R8'),('S8','T8'),('U8','V8'),
    ('X8','Y8'),('Z8','AA8'),('AB8','AC8'),('AD8','AE8'),('AF8','AG8'),('AH8','AI8'),
    ('AK8','AL8'),('AM8','AN8'),('AO8','AP8'),('AQ8','AR8'),('AS8','AT8'),

    ('K7','K8'),('L7','L8'),('W7','W8'),('AJ7','AJ8')
]
# ============================================================================
# CLASES DE UTILIDAD PARA ESTILOS
# ============================================================================

class ExcelStyleManager:
    """Gestor centralizado de estilos para Excel."""
    
    _fills_cache = {}
    _fonts_cache = {}
    _borders_cache = {}
    
    @classmethod
    def get_fill(cls, color_key):
        """Obtiene un PatternFill cacheado."""
        if color_key not in cls._fills_cache:
            color = COLORS.get(color_key, color_key)
            cls._fills_cache[color_key] = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )
        return cls._fills_cache[color_key]
    
    @classmethod
    def get_font(cls, name='Arial', size=8, bold=False, color='000000'):
        """Obtiene una Font cacheada."""
        key = (name, size, bold, color)
        if key not in cls._fonts_cache:
            cls._fonts_cache[key] = Font(name=name, size=size, bold=bold, color=color)
        return cls._fonts_cache[key]
    
    @classmethod
    def get_border(cls, color='A9A9A9', style='thin'):
        """Obtiene un Border cacheado."""
        key = (color, style)
        if key not in cls._borders_cache:
            side = Side(style=style, color=color)
            cls._borders_cache[key] = Border(
                left=side, right=side, top=side, bottom=side
            )
        return cls._borders_cache[key]
    
    @classmethod
    def get_alignment(cls, horizontal='center', vertical='center', wrap_text=False):
        """Obtiene un Alignment."""
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


# ============================================================================
# CLASE BASE PARA REPORTES
# ============================================================================

class BaseExcelReportView(LoginRequiredMixin, View):
    """Clase base para generar reportes Excel."""
    
    filename = "reporte.xlsx"
    sheet_name = "Datos"
    
    def get_query_params(self, request):
        """Extrae los parámetros de consulta. Debe ser implementado."""
        raise NotImplementedError("Subclases deben implementar get_query_params()")
    
    def get_data(self, params):
        """Obtiene los datos del reporte. Debe ser implementado."""
        raise NotImplementedError("Subclases deben implementar get_data()")
    
    def get_filename(self):
        """Retorna el nombre del archivo."""
        return self.filename
    
    def get(self, request, *args, **kwargs):
        """Maneja la petición GET y genera el Excel."""
        params = self.get_query_params(request)
        data = self.get_data(params)
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        
        fill_worksheet_optimized(ws, data, request.user)
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={self.get_filename()}"
        wb.save(response)
        
        return response


# ============================================================================
# VISTAS DE REPORTES
# ============================================================================

class RptPaqueteGestante(BaseExcelReportView):
    """Reporte de captación de gestantes."""
    
    filename = "rpt_dashboard.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
            'provincia': request.GET.get('provincia', ''),
            'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('p_microredes', ''),
            'establecimiento': request.GET.get('p_establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_dashboard(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            params['provincia'], params['distrito'], params['red'],
            params['microredes'], params['establecimiento'], params['cumple']
        )


class RptPaqueteGestanteMicroRed(BaseExcelReportView):
    """Reporte de población por microred."""
    
    filename = "rpt_dashboard_microred.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
           # 'provincia': request.GET.get('provincia', ''),
           # 'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('p_microredes', ''),
           # 'establecimiento': request.GET.get('p_establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_dashboard(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            '',  # provincia (vacío para microred)
            '',  # distrito (vacío para microred)
            params['red'],params['microredes'], 
              '',  # establecimiento (vacío para microred)
            params['cumple']
        )


class RptPaqueteGestanteEstablec(BaseExcelReportView):
    """Reporte de población por establecimiento."""
    
    filename = "rpt_dashboard_establecimiento.xlsx"
    sheet_name = "Seguimiento"
    
    def get_query_params(self, request):
        return {
            'anio': request.GET.get('anio', '2025'),
            'mes_inicio': request.GET.get('fecha_inicio', ''),
            'mes_fin': request.GET.get('fecha_fin', ''),
            'provincia': request.GET.get('provincia', ''),
            'distrito': request.GET.get('distrito', ''),
            'red': request.GET.get('red', ''),
            'microredes': request.GET.get('microred', ''),
            'establecimiento': request.GET.get('establecimiento', ''),
            'cumple': request.GET.get('cumple', ''),
        }
    
    def get_data(self, params):
        return obtener_seguimiento_dashboard(
            params['anio'], params['mes_inicio'], params['mes_fin'],
            params['provincia'], params['distrito'], params['red'],
            params['microredes'], params['establecimiento'], params['cumple']
        )


# ============================================================================
# FUNCIONES DE FORMATEO
# ============================================================================

def fill_worksheet_optimized(ws, results, user=None):
    """Función optimizada para llenar la hoja de trabajo."""
    
    style_mgr = ExcelStyleManager
    
    # Configurar dimensiones
    _set_dimensions(ws)
    
    # Configurar agrupación de columnas
    _setup_column_grouping(ws)
    
    # Congelar paneles
    # ws.freeze_panes = 'S10'
    
    # Configurar celdas combinadas
    _setup_merged_cells(ws)
    
    # Aplicar estilos a secciones
    _style_header_sections(ws, style_mgr)
    
    # Configurar cabeceras de columnas
    _setup_column_headers(ws, style_mgr)
    
    # Agregar metadatos del reporte
    _add_report_metadata(ws, user, style_mgr)
    
    # Agregar títulos
    _add_titles(ws, style_mgr)
    
    # Escribir datos
    _write_data(ws, results, style_mgr)


def _set_dimensions(ws):
    """Configura las dimensiones de filas y columnas."""
    for row, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height
    
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _setup_column_grouping(ws):
    """Configura el agrupamiento de columnas."""
    #grouped_columns = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    #ws.column_dimensions.group('K', 'R', hidden=True)
    
    #for col in grouped_columns:
    #    ws.column_dimensions[col].width = COLUMN_WIDTHS.get(col, 10)
    #    ws.column_dimensions[col].outline_level = 1


def _setup_merged_cells(ws):
    """Configura las celdas combinadas."""
    for start, end in MERGE_CELLS_CONFIG:
        ws.merge_cells(f'{start}:{end}')


def _style_header_sections(ws, style_mgr):
    """Aplica estilos a las secciones de cabecera."""
    
    border_negro = style_mgr.get_border('000000')
    
    # Configuración de secciones con sus textos y estilos
    # NOTA: Solo incluir la celda INICIAL de cada rango combinado (MergedCell)
    # Las celdas dentro de un rango combinado que no son la inicial son read-only
    sections_config = {
        # Fila 5: B5:G5 y H5:K5 son rangos combinados, L5 es celda individual
        'B5': ('META (DENOMINADOR)', 'gray', 10, True),
        'L5': ('NUMERADOR (N° de mujeres del denominador que durante su gestación,recibieron el paquete integrado de servicios)', 'naranja_claro', 10, True),
        # Fila 6: B6:G6, H6:K6 son rangos combinados, L6 es celda individual
        'B6': ('N° de mujeres procedentes de los distritos de quintiles 1 y 2 de pobreza departamental con parto institucional, según la base de datos del CNV en línea', 'gray', 10, True),
        'L6': ('Recibieron en el primer trimestre: 4 examenes auxiliares o Perfil Obstetrico', 'gray', 8, True),
        'W6': ('Atencion Prenatal', 'gray', 8, True),
        'X6': ('APN en el PRIMER TRIMESTRE', 'gray', 8, True),
        'Z6': ('APN en el SEGUNDO TRIMESTRE', 'gray', 8, True),
        'AD6': ('APN en el TERCER TRIMESTRE', 'gray', 8, True),
        'AJ6': ('Entrega de sulfato ferroso + acido folico', 'gray', 8, True),
        # Fila 7: B7:D7, E7:H7, I7:J7, N7:P7, Q7:R7, T7:V7, W7:X7 son rangos combinados
        # NOTA: K7, L7 NO están combinadas (son celdas individuales)
        'B7': ('INTEVALOS', 'plomo_claro', 7, True),
        'D7': ('CNV EN LINEA', 'plomo_claro', 7, False),
        'E7': ('FUR Calculado', 'plomo_claro', 7, False),
        'F7': ('98 dias desde el FUR', 'plomo_claro', 7, False),
        'G7': ('196 dias desde el FUR', 'plomo_claro', 7, False),
        'H7': ('231 dias desde el FUR', 'plomo_claro', 7, False),
        'I7': ('259 dias desde el FUR', 'plomo_claro', 7, False),
        'J7': ('Nacimiento del CNV en Linea', 'plomo_claro', 7, False),
        'K7': ('DENOMINADOR', 'plomo_claro', 7, False),
        'L7': ('NUMERADOR PARCIAL', 'plomo_claro', 7, False),
        'M7': ('Tamizaje de hemoglobina', 'plomo_claro', 7, False),
        'O7': ('Tamizaje de sifilis con prueba rapida', 'plomo_claro', 7, False),
        'Q7': ('Tamizaje de VIH con prueba rapida', 'plomo_claro', 7, False),
        'S7': ('Tamizaje de bacteriuria asintomática', 'plomo_claro', 7, False),
        'U7': ('Un examen de perfil obstetrico', 'plomo_claro', 7, False),
        'W7': ('NUMERADOR PARCIAL', 'plomo_claro', 7, False),
        'X7': ('Al menos 1 APN (< semana 14)', 'plomo_claro', 7, False),
        'Z7': ('Al menos 1 APN (semana 14 a < 28), Intervalo Minimo 25 dias des APN anterior', 'plomo_claro', 7, False),
        'AB7': ('Al menos 2 APN (semana 14 < 28 ), Intervalo Minimo 25 dias des APN anterior', 'plomo_claro', 7, False),
        'AD7': ('Al menos 1° APN (semana 28 a parto)', 'plomo_claro', 7, False),
        'AF7': ('Al menos 2° APN (semana 28 a parto)', 'plomo_claro', 7, False),
        'AH7': ('Al menos 3° APN (semana 28 a parto)', 'plomo_claro', 7, False),
        'AJ7': ('NUMERADOR PARCIAL', 'plomo_claro', 7, False),
        'AK7': ('Al menos 1 entrega de sulfato ferroso + Acido Folico', 'plomo_claro', 7, False),
        'AM7': ('Al menos 2° Entrega (25-40 dias desde 1ra)', 'plomo_claro', 7, False),
        'AO7': ('Al menos 3° Entrega (25-40 dias desde 2da)', 'plomo_claro', 7, False),
        'AQ7': ('Al menos 4° Entrega (25-40 dias desde 3ra)', 'plomo_claro', 7, False),
        'AS7': ('Al menos 5° Entrega (25-40 dias desde 4ta)', 'plomo_claro', 7, False),





        # Fila 8: B8:D8, E8:H8, I8:J8, N8:P8, Q8:R8, T8:V8, W8:X8 son rangos combinados
        # NOTA: K8, L8, M8, S8, Y8, Z8, AA8 NO están combinadas
        'B8': ('CODIGO HIS MINSA', 'azul_claro', 7, True),
        'D8': ('>= 37 sem', 'azul_claro', 7, False),
        'E8': ('PARTO - (SG*7) = -273 dias', 'azul_claro', 7, False),
        'F8': ('Limite 1° TRIM', 'azul_claro', 7, False),
        'G8': ('Limite 2° TRIM', 'azul_claro', 7, False),
        'H8': ('APN (25 a 13 dias)', 'azul_claro', 7, False),
        'I8': ('APN (13 a 7)', 'azul_claro', 7, False),
        'J8': ('Delimita periodo gestacional', 'azul_claro', 7, False),
        'M8': ('DX=85018 ó 85018.01 ó 85031', 'azul_claro', 7, False),
        'O8': ('DX=86780 ó 86592 ó 86593 ó 86318.01 ó 86780.01', 'azul_claro', 7, False),
        'Q8': ('DX=86703 ó 86703.02 ó 87389 ó 86318.01 ó 86703.01', 'azul_claro', 7, False),
        'S8': ('DX = 81007 ó 81002 ó 81000.02', 'azul_claro', 7, False),
        'U8': ('DX= 80055.01', 'azul_claro', 7, False),
        'X8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'Z8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'AB8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'AD8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'AF8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'AH8': ('DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593', 'azul_claro', 7, False),
        'AK8': ('DX= 99199.26 ó (DX= 99199.26 + DX = O990 ó D509)', 'azul_claro', 7, False),
        'AM8': ('DX= 99199.26 ó (DX= 99199.26 + DX = O990 ó D509)', 'azul_claro', 7, False),
        'AO8': ('DX= 99199.26 ó (DX= 99199.26 + DX = O990 ó D509)', 'azul_claro', 7, False),
        'AQ8': ('DX= 99199.26 ó (DX= 99199.26 + DX = O990 ó D509)', 'azul_claro', 7, False),
        'AS8': ('DX= 99199.26 ó (DX= 99199.26 + DX = O990 ó D509)', 'azul_claro', 7, False),
    }
    
    for cell_ref, (text, fill_color, font_size, bold) in sections_config.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(wrap_text=True)
        cell.font = style_mgr.get_font(size=font_size, bold=bold)
        cell.fill = style_mgr.get_fill(fill_color)
        cell.border = border_negro
    
    # Aplicar bordes a las filas de cabecera
    _apply_row_borders(ws, [5, 6, 7, 8], 'B', 'AT', border_negro)


def _apply_row_borders(ws, rows, start_col, end_col, border):
    """Aplica bordes a rangos de celdas."""
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    
    for row in rows:
        for col in range(start_idx, end_idx + 1):
            ws.cell(row=row, column=col).border = border


def _setup_column_headers(ws, style_mgr):
    """Configura las cabeceras de columnas."""
    
    border = style_mgr.get_border('00B0F0')
    
    for cell_ref, text, fill_color in HEADERS_CONFIG:
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(wrap_text=True)
        cell.font = style_mgr.get_font(size=8, bold=True)
        cell.fill = style_mgr.get_fill(fill_color)
        cell.border = border


def _add_report_metadata(ws, user, style_mgr):
    """Agrega metadatos del reporte (fecha, hora, usuario)."""
    
    fecha_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = user.username if user else getpass.getuser()
    
    metadata = [
        ('Q1', 'Fecha y Hora:', 'R1', fecha_hora),
        ('Q2', 'Usuario:', 'R2', nombre_usuario),
    ]
    
    etiqueta_font = style_mgr.get_font(size=8)
    
    for label_ref, label, value_ref, value in metadata:
        ws[label_ref].value = label
        ws[label_ref].font = etiqueta_font
        ws[label_ref].alignment = style_mgr.get_alignment(horizontal='right')
        
        ws[value_ref].value = value
        ws[value_ref].font = etiqueta_font
        ws[value_ref].alignment = style_mgr.get_alignment(horizontal='left')


def _add_titles(ws, style_mgr):
    """Agrega los títulos del reporte."""
    
    titles = [
        ('B1', 'OFICINA DE TECNOLOGIAS DE LA INFORMACION', 7, True, '000000'),
        ('B2', 'DIRECCION REGIONAL DE SALUD JUNIN', 7, True, '000000'),
        ('B3', 'El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.', 7, True, '0000CC'),
        ('B4', 'SEGUIMIENTO NOMINAL:MC-01. MUJERES CON PARTO INSTITUCIONAL, PROCEDENTES DE LOS DISTRITOS DE QUINTILES 1 Y 2 DE POBREZA DEPARTAMENTAL, QUE DURANTE SU GESTACIÓN RECIBIERON EL PAQUETE INTEGRADO DE SERVICIOS', 12, True, '000000'),
    ]
    
    for cell_ref, text, size, bold, color in titles:
        cell = ws[cell_ref]
        cell.value = text
        cell.alignment = style_mgr.get_alignment(horizontal='left')
        cell.font = style_mgr.get_font(size=size, bold=bold, color=color)


def _write_data(ws, results, style_mgr):
    """Escribe los datos en la hoja de trabajo."""
    
    border = style_mgr.get_border()
    check_mark = '✓'
    x_mark = '✗'
    
    # Columnas con alineación izquierda
    left_align_cols = {52, 54}
    # Columnas con check/x marks
    check_cols = {14,16,18,20,22,25,27,29,31,33, 35,38,40,42,44,46}
    # Columnas con check/x marks validador
    check_cols_validador = {11}
    # Columnas de sub-indicadores
    sub_indicator_cols = {12,23,36}
    
    for row_idx, record in enumerate(results, start=10):
        for col_idx, value in enumerate(record.values(), start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Determinar alineación
            if col_idx in left_align_cols:
                cell.alignment = style_mgr.get_alignment(horizontal='left')
            else:
                cell.alignment = style_mgr.get_alignment()
            
            # Aplicar formato según columna
            if col_idx == 47:  # Columna INDICADOR
                _format_indicator_cell(cell, value, style_mgr)
            elif col_idx in check_cols:
                _format_check_cell(cell, value, check_mark, x_mark, style_mgr)
            elif col_idx in check_cols_validador:
                _format_check_cell_validador(cell, value, check_mark, x_mark, style_mgr)
            elif col_idx in sub_indicator_cols:
                _format_sub_indicator_cell(cell, value, style_mgr)
            else:
                cell.font = style_mgr.get_font(size=8)


def _format_indicator_cell(cell, value, style_mgr):
    """Formatea la celda de indicador."""
    if value == 0:
        cell.value = 'NO CUMPLE'
        cell.fill = style_mgr.get_fill('red')
        cell.font = style_mgr.get_font(size=8, bold=True, color='000000')
    elif value == 1:
        cell.value = 'CUMPLE'
        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')
        cell.font = style_mgr.get_font(size=8, bold=True, color='000000')
    else:
        cell.font = style_mgr.get_font(size=8, bold=True)


def _format_check_cell(cell, value, check_mark, x_mark, style_mgr):
    """Formatea celdas con check/x marks."""
    if value == 1:
        cell.value = check_mark
        cell.font = style_mgr.get_font(size=10, color='00B050')
    elif value == 0:
        cell.value = x_mark
        cell.font = style_mgr.get_font(size=10, color='FF0000')
    else:
        cell.font = style_mgr.get_font(size=8)

def _format_check_cell_validador(cell, value, check_mark, x_mark, style_mgr):
    """Formatea celdas con check/x marks."""
    if value == 1:
        cell.value = check_mark
        cell.font = style_mgr.get_font(size=10, color='00B050')
        cell.fill = style_mgr.get_fill('green')
    elif value == 0:
        cell.value = x_mark
        cell.font = style_mgr.get_font(size=10, color='FF0000')
    else:
        cell.font = style_mgr.get_font(size=8)



def _format_sub_indicator_cell(cell, value, style_mgr):
    """Formatea celdas de sub-indicadores."""
    if value == 0:
        cell.value = 'NO CUMPLE'
        cell.font = style_mgr.get_font(size=7, color='FF0000')
        cell.fill = style_mgr.get_fill('gray')
    elif value == 1:
        cell.value = 'CUMPLE'
        cell.font = style_mgr.get_font(size=7, color='00B050')
        cell.fill = style_mgr.get_fill('gray')  
    else:
        cell.font = style_mgr.get_font(size=7)